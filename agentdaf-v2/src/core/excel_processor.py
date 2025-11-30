"""
Excel Processor - Advanced Excel file processing and data extraction
"""

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelProcessor:
    def __init__(self):
        self.supported_formats = [".xlsx", ".xls", ".csv"]
        self.data_cache = {}

    def is_supported_file(self, file_path: str) -> bool:
        """Check if file format is supported"""
        return Path(file_path).suffix.lower() in self.supported_formats

    def load_excel_file(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Load Excel file and return data information"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")

            if not self.is_supported_file(file_path):
                raise ValueError(f"Unsupported file format: {Path(file_path).suffix}")

            file_info = {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "file_size": os.path.getsize(file_path),
                "modified_time": datetime.fromtimestamp(os.path.getmtime(file_path)),
                "sheets": [],
                "data": {},
            }

            # Load based on file type
            if file_path.endswith(".csv"):
                df = pd.read_csv(file_path)
                file_info["sheets"].append("Sheet1")
                file_info["data"]["Sheet1"] = df
            else:
                # Excel file
                wb = load_workbook(file_path, read_only=True)
                file_info["sheets"] = wb.sheetnames

                for sheet in wb.sheetnames:
                    if sheet_name and sheet != sheet_name:
                        continue

                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet)
                        file_info["data"][sheet] = df
                    except Exception as e:
                        logger.warning(f"Error loading sheet {sheet}: {e}")
                        continue

                wb.close()

            # Cache the data
            self.data_cache[file_path] = file_info

            logger.info(
                f"Loaded Excel file: {file_path} with {len(file_info['sheets'])} sheets"
            )
            return file_info

        except Exception as e:
            logger.error(f"Error loading Excel file {file_path}: {e}")
            raise

    def get_sheet_info(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """Get detailed information about a specific sheet"""
        try:
            file_info = self.load_excel_file(file_path)

            if sheet_name not in file_info["data"]:
                raise ValueError(f"Sheet '{sheet_name}' not found in file")

            df = file_info["data"][sheet_name]

            sheet_info = {
                "sheet_name": sheet_name,
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": list(df.columns),
                "data_types": df.dtypes.to_dict(),
                "null_counts": df.isnull().sum().to_dict(),
                "memory_usage": df.memory_usage(deep=True).sum(),
                "sample_data": df.head(5).to_dict("records") if len(df) > 0 else [],
            }

            # Add basic statistics for numeric columns
            numeric_columns = df.select_dtypes(include=[np.number]).columns
            if len(numeric_columns) > 0:
                sheet_info["statistics"] = df[numeric_columns].describe().to_dict()

            return sheet_info

        except Exception as e:
            logger.error(f"Error getting sheet info for {sheet_name}: {e}")
            raise

    def clean_data(
        self, df: pd.DataFrame, options: Dict[str, Any] = None
    ) -> pd.DataFrame:
        """Clean and preprocess DataFrame"""
        try:
            if options is None:
                options = {}

            cleaned_df = df.copy()

            # Remove empty rows and columns
            if options.get("remove_empty_rows", True):
                cleaned_df = cleaned_df.dropna(how="all")

            if options.get("remove_empty_columns", True):
                cleaned_df = cleaned_df.dropna(axis=1, how="all")

            # Fill missing values
            fill_method = options.get("fill_method", "forward")
            if fill_method == "forward":
                cleaned_df = cleaned_df.fillna(method="ffill")
            elif fill_method == "backward":
                cleaned_df = cleaned_df.fillna(method="bfill")
            elif fill_method == "mean":
                numeric_columns = cleaned_df.select_dtypes(include=[np.number]).columns
                cleaned_df[numeric_columns] = cleaned_df[numeric_columns].fillna(
                    cleaned_df[numeric_columns].mean()
                )
            elif fill_method == "zero":
                cleaned_df = cleaned_df.fillna(0)

            # Remove duplicates
            if options.get("remove_duplicates", True):
                cleaned_df = cleaned_df.drop_duplicates()

            # Convert data types
            if options.get("auto_convert_types", True):
                cleaned_df = self._auto_convert_types(cleaned_df)

            logger.info(f"Data cleaned: {len(df)} -> {len(cleaned_df)} rows")
            return cleaned_df

        except Exception as e:
            logger.error(f"Error cleaning data: {e}")
            raise

    def _auto_convert_types(self, df: pd.DataFrame) -> pd.DataFrame:
        """Automatically convert data types"""
        try:
            # Convert numeric columns
            for col in df.columns:
                if df[col].dtype == "object":
                    # Try to convert to numeric
                    try:
                        df[col] = pd.to_numeric(df[col], errors="ignore")
                    except:
                        pass

                    # Try to convert to datetime
                    if df[col].dtype == "object":
                        try:
                            df[col] = pd.to_datetime(df[col], errors="ignore")
                        except:
                            pass

            return df

        except Exception as e:
            logger.error(f"Error auto-converting types: {e}")
            return df

    def analyze_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Perform comprehensive data analysis"""
        try:
            analysis = {
                "basic_info": {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "memory_usage_mb": df.memory_usage(deep=True).sum() / 1024 / 1024,
                },
                "column_analysis": {},
                "data_quality": {},
                "correlations": {},
            }

            # Analyze each column
            for col in df.columns:
                col_analysis = {
                    "data_type": str(df[col].dtype),
                    "null_count": df[col].isnull().sum(),
                    "null_percentage": (df[col].isnull().sum() / len(df)) * 100,
                    "unique_count": df[col].nunique(),
                    "unique_percentage": (df[col].nunique() / len(df)) * 100,
                }

                # Numeric analysis
                if pd.api.types.is_numeric_dtype(df[col]):
                    col_analysis.update(
                        {
                            "min": df[col].min(),
                            "max": df[col].max(),
                            "mean": df[col].mean(),
                            "median": df[col].median(),
                            "std": df[col].std(),
                            "skewness": df[col].skew(),
                            "kurtosis": df[col].kurtosis(),
                        }
                    )

                # Text analysis
                elif df[col].dtype == "object":
                    col_analysis.update(
                        {
                            "avg_length": df[col].astype(str).str.len().mean(),
                            "max_length": df[col].astype(str).str.len().max(),
                            "min_length": df[col].astype(str).str.len().min(),
                        }
                    )

                analysis["column_analysis"][col] = col_analysis

            # Data quality metrics
            analysis["data_quality"] = {
                "completeness_score": (
                    1 - df.isnull().sum().sum() / (len(df) * len(df.columns))
                )
                * 100,
                "duplicate_rows": df.duplicated().sum(),
                "duplicate_percentage": (df.duplicated().sum() / len(df)) * 100,
            }

            # Correlation analysis for numeric columns
            numeric_df = df.select_dtypes(include=[np.number])
            if len(numeric_df.columns) > 1:
                analysis["correlations"] = numeric_df.corr().to_dict()

            return analysis

        except Exception as e:
            logger.error(f"Error analyzing data: {e}")
            raise

    def create_dashboard_data(
        self, df: pd.DataFrame, config: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """Create data specifically for dashboard generation"""
        try:
            if config is None:
                config = {}

            dashboard_data = {"summary": {}, "charts": {}, "tables": {}}

            # Summary statistics
            dashboard_data["summary"] = {
                "total_records": len(df),
                "total_columns": len(df.columns),
                "last_updated": datetime.now().isoformat(),
            }

            # Generate chart data
            chart_types = config.get("chart_types", ["bar", "line", "pie"])

            for chart_type in chart_types:
                if chart_type == "bar" and len(df.columns) > 0:
                    # Bar chart for first categorical column
                    cat_cols = df.select_dtypes(include=["object"]).columns
                    if len(cat_cols) > 0:
                        col = cat_cols[0]
                        value_counts = df[col].value_counts().head(10)
                        dashboard_data["charts"][f"{col}_bar"] = {
                            "type": "bar",
                            "data": {
                                "labels": value_counts.index.tolist(),
                                "values": value_counts.values.tolist(),
                            },
                        }

                elif chart_type == "line" and len(df.columns) > 0:
                    # Line chart for first numeric column
                    num_cols = df.select_dtypes(include=[np.number]).columns
                    if len(num_cols) > 0:
                        col = num_cols[0]
                        dashboard_data["charts"][f"{col}_line"] = {
                            "type": "line",
                            "data": {
                                "labels": list(range(min(100, len(df)))),
                                "values": df[col].head(100).tolist(),
                            },
                        }

            # Table data
            dashboard_data["tables"]["main_data"] = {
                "columns": df.columns.tolist(),
                "data": df.head(100).to_dict("records"),
            }

            return dashboard_data

        except Exception as e:
            logger.error(f"Error creating dashboard data: {e}")
            raise

    def export_to_excel(
        self,
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        output_path: str,
        include_analysis: bool = False,
    ) -> bool:
        """Export data to Excel file"""
        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                if isinstance(data, pd.DataFrame):
                    # Single DataFrame
                    data.to_excel(writer, sheet_name="Data", index=False)

                    if include_analysis:
                        analysis = self.analyze_data(data)
                        analysis_df = pd.DataFrame([analysis])
                        analysis_df.to_excel(writer, sheet_name="Analysis", index=False)

                elif isinstance(data, dict):
                    # Multiple DataFrames
                    for sheet_name, df in data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Apply formatting
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]

                    # Format header row
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                        )
                        cell.alignment = Alignment(horizontal="center")

                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = (
                            adjusted_width
                        )

            logger.info(f"Data exported to {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False

    def get_file_summary(self, file_path: str) -> Dict[str, Any]:
        """Get quick summary of Excel file"""
        try:
            file_info = self.load_excel_file(file_path)

            summary = {
                "file_name": file_info["file_name"],
                "file_size_mb": file_info["file_size"] / 1024 / 1024,
                "modified_time": file_info["modified_time"].isoformat(),
                "sheet_count": len(file_info["sheets"]),
                "sheets": {},
            }

            for sheet_name in file_info["sheets"]:
                if sheet_name in file_info["data"]:
                    df = file_info["data"][sheet_name]
                    summary["sheets"][sheet_name] = {
                        "rows": len(df),
                        "columns": len(df.columns),
                        "has_data": len(df) > 0,
                    }

            return summary

        except Exception as e:
            logger.error(f"Error getting file summary: {e}")
            raise


# Global Excel processor instance
excel_processor = ExcelProcessor()


# Utility functions
def process_gaming_data(file_path: str) -> Dict[str, Any]:
    """Process gaming data Excel file specifically"""
    try:
        processor = ExcelProcessor()
        file_info = processor.load_excel_file(file_path)

        # Look for gaming-specific sheets
        gaming_sheets = ["players", "alliances", "matches", "statistics"]
        processed_data = {}

        for sheet_name in file_info["sheets"]:
            df = file_info["data"][sheet_name]

            # Clean and analyze data
            cleaned_df = processor.clean_data(df)
            analysis = processor.analyze_data(cleaned_df)

            processed_data[sheet_name] = {
                "data": cleaned_df,
                "analysis": analysis,
                "dashboard_data": processor.create_dashboard_data(cleaned_df),
            }

        return {
            "file_info": file_info,
            "processed_data": processed_data,
            "summary": processor.get_file_summary(file_path),
        }

    except Exception as e:
        logger.error(f"Error processing gaming data: {e}")
        raise

    # Backwards-compatibility helper methods expected by API/tests
    def process_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Process an Excel or CSV file and return structured data"""
        try:
            file_info = self.load_excel_file(file_path)
            processed = {"file_info": file_info, "processed": {}}
            for sheet_name in file_info["sheets"]:
                df = file_info["data"].get(sheet_name)
                if isinstance(df, pd.DataFrame):
                    cleaned = self.clean_data(df)
                    analysis = self.analyze_data(cleaned)
                    processed["processed"][sheet_name] = {
                        "data": cleaned.to_dict("records"),
                        "analysis": analysis,
                        "dashboard_data": self.create_dashboard_data(cleaned),
                    }
            return processed
        except Exception as e:
            logger.error(f"Error in process_excel_file: {e}")
            raise

    def validate_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate incoming structured data or file-info dict and return validation dict"""
        try:
            errors = []
            if not data or "sheets" not in data:
                errors.append("No sheets found")
                return {"is_valid": False, "errors": errors}

            for sheet_name, sheet in data.get("sheets", {}).items():
                if not sheet or "data" not in sheet or not sheet.get("data"):
                    errors.append(f"No data in sheet {sheet_name}")
            return {"is_valid": len(errors) == 0, "errors": errors}
        except Exception as e:
            logger.error(f"Validation error: {e}")
            return {"is_valid": False, "errors": [str(e)]}

    def extract_player_data(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract player records from processed data dict"""
        players = []
        if not data:
            return players

        # Support two shapes: {'data': [rows]} or {'sheets': {}} style
        if "data" in data and isinstance(data["data"], list):
            for row in data["data"]:
                if isinstance(row, dict) and ("player" in row or "name" in row):
                    name = row.get("player") or row.get("name")
                    players.append(
                        {
                            "name": name,
                            "score": row.get("score", 0),
                            "alliance": row.get("alliance", None),
                        }
                    )
        elif "sheets" in data:
            for sheet_name, sheet in data.get("sheets", {}).items():
                if isinstance(sheet, dict) and "data" in sheet:
                    for row in sheet["data"]:
                        if isinstance(row, dict) and ("player" in row or "name" in row):
                            name = row.get("player") or row.get("name")
                            players.append(
                                {
                                    "name": name,
                                    "score": row.get("score", 0),
                                    "alliance": row.get("alliance", None),
                                }
                            )
        return players

    def process_game_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Process structured game data (dict) and generate a result expected by tests"""
        try:
            if not data:
                return {
                    "summary": {"total_games": 0, "total_players": 0},
                    "games": [],
                    "charts": {},
                }
            total_games = len(data.get("sheets", {})) if "sheets" in data else 0
            players = self.extract_player_data(data)
            total_players = len(players)
            # Build charts placeholder (simple statistics)
            charts = {}
            games = []
            if "sheets" in data:
                for sheet_name, sheet in data.get("sheets", {}).items():
                    games.append(
                        {
                            "sheet_name": sheet_name,
                            "total_players": len(sheet.get("data", [])),
                            "stats": (
                                sheet.get("statistics", {})
                                if isinstance(sheet, dict)
                                else {}
                            ),
                        }
                    )

            summary = {"total_games": total_games, "total_players": total_players}
            return {"summary": summary, "games": games, "charts": charts}
        except Exception as e:
            logger.error(f"Error in process_game_data: {e}")
            return {
                "summary": {"total_games": 0, "total_players": 0},
                "games": [],
                "charts": {},
            }
